import io
from datetime import datetime, timedelta

import openpyxl

from app import db
from app.models import MaintenanceTask, ReportLog, WaterPoint
from app.report_queries import bucket_risk_level
from tests.conftest import login, make_user, make_water_point


def _seed_basic_fleet(district="Bugesera"):
    wp_functional = make_water_point(db, water_point_id="WP-F1", district=district)
    wp_functional.current_status = "Functional"
    wp_functional.risk_probability = 0.1
    wp_functional.last_prediction_date = datetime.utcnow()

    wp_at_risk = make_water_point(db, water_point_id="WP-R1", district=district)
    wp_at_risk.current_status = "At Risk"
    wp_at_risk.risk_probability = 0.5
    wp_at_risk.last_prediction_date = datetime.utcnow()

    wp_repair = make_water_point(db, water_point_id="WP-U1", district=district)
    wp_repair.current_status = "Under Repair"
    wp_repair.risk_probability = 0.8
    wp_repair.last_prediction_date = datetime.utcnow()

    wp_unassessed = make_water_point(db, water_point_id="WP-N1", district=district)
    wp_unassessed.current_status = "Non-Functional"
    wp_unassessed.risk_probability = 0.0
    wp_unassessed.last_prediction_date = None

    db.session.commit()
    return wp_functional, wp_at_risk, wp_repair, wp_unassessed


def _seed_completed_task(manager, tech, wp, hours_to_resolve=4):
    now = datetime.utcnow()
    task = MaintenanceTask(
        water_point_id=wp.id,
        created_by_id=manager.id,
        assigned_to_id=tech.id,
        title="Repair leak",
        status="completed",
        assigned_at=now - timedelta(hours=hours_to_resolve),
        completed_at=now,
        resulting_status="Functional",
        completion_notes="Replaced gasket",
    )
    db.session.add(task)
    db.session.commit()
    return task


# --- RBAC ---------------------------------------------------------------

REPORT_ROUTES = [
    "/reports/status",
    "/reports/technician-performance",
    "/reports/maintenance",
    "/reports/predictive-risk",
    "/reports/district-summary",
]


def test_unauthenticated_redirected_to_login(client):
    for route in REPORT_ROUTES:
        resp = client.get(route)
        assert resp.status_code == 302
        assert "/auth/login" in resp.headers["Location"]


def test_all_roles_can_view_reports(app, client):
    with app.app_context():
        make_user(db, "viewer", "Bugesera", "viewer1")
        make_user(db, "district_manager", "Bugesera", "manager1")
        make_user(db, "district_technician", "Bugesera", "tech1")
        make_user(db, "admin", "Bugesera", "admin1")
        _seed_basic_fleet()

    for username in ("viewer1", "manager1", "tech1", "admin1"):
        login(client, username)
        for route in REPORT_ROUTES:
            resp = client.get(route)
            assert resp.status_code == 200, f"{username} could not view {route}"
        client.get("/auth/logout")


def test_maintenance_report_is_district_scoped_for_non_admin(app, client):
    with app.app_context():
        manager_a = make_user(db, "district_manager", "Bugesera", "managerA")
        tech_a = make_user(db, "district_technician", "Bugesera", "techA")
        wp_a = make_water_point(db, water_point_id="WP-A1", district="Bugesera")
        _seed_completed_task(manager_a, tech_a, wp_a)

        manager_b = make_user(db, "district_manager", "Nyagatare", "managerB")
        tech_b = make_user(db, "district_technician", "Nyagatare", "techB")
        wp_b = make_water_point(db, water_point_id="WP-B1", district="Nyagatare")
        _seed_completed_task(manager_b, tech_b, wp_b)

    login(client, "managerA")
    resp = client.get("/reports/maintenance")
    assert b"WP-A1" in resp.data
    assert b"WP-B1" not in resp.data


def test_technician_only_sees_own_row_in_performance_report(app, client):
    with app.app_context():
        manager = make_user(db, "district_manager", "Bugesera", "manager1")
        tech1 = make_user(db, "district_technician", "Bugesera", "tech1")
        tech2 = make_user(db, "district_technician", "Bugesera", "tech2")
        wp = make_water_point(db, water_point_id="WP-1", district="Bugesera")
        _seed_completed_task(manager, tech1, wp)
        wp2 = make_water_point(db, water_point_id="WP-2", district="Bugesera")
        _seed_completed_task(manager, tech2, wp2)

    login(client, "tech1")
    resp = client.get("/reports/technician-performance")
    assert b"Tech1" in resp.data
    assert b"Tech2" not in resp.data


# --- Data correctness -----------------------------------------------------


def test_status_report_summary_counts(app, client):
    with app.app_context():
        make_user(db, "admin", "Bugesera", "admin1")
        _seed_basic_fleet()

    login(client, "admin1")
    resp = client.get("/reports/status")
    html = resp.data.decode()
    assert resp.status_code == 200
    # one of each status seeded
    assert ">4<" in html  # total


def test_predictive_risk_report_excludes_unassessed_points(app, client):
    with app.app_context():
        make_user(db, "admin", "Bugesera", "admin1")
        _seed_basic_fleet()

    login(client, "admin1")
    resp = client.get("/reports/predictive-risk")
    assert b"WP-N1" not in resp.data  # last_prediction_date is None
    assert b"WP-F1" in resp.data


def test_risk_level_bucket_thresholds():
    assert bucket_risk_level(0.0) == "Low"
    assert bucket_risk_level(0.32) == "Low"
    assert bucket_risk_level(0.33) == "Medium"
    assert bucket_risk_level(0.65) == "Medium"
    assert bucket_risk_level(0.66) == "High"
    assert bucket_risk_level(0.99) == "High"
    assert bucket_risk_level(None) is None


def test_technician_performance_completion_rate_and_resolution_time(app, client):
    with app.app_context():
        manager = make_user(db, "district_manager", "Bugesera", "manager1")
        tech = make_user(db, "district_technician", "Bugesera", "tech1")
        wp1 = make_water_point(db, water_point_id="WP-1", district="Bugesera")
        wp2 = make_water_point(db, water_point_id="WP-2", district="Bugesera")
        _seed_completed_task(manager, tech, wp1, hours_to_resolve=2)
        _seed_completed_task(manager, tech, wp2, hours_to_resolve=6)

    login(client, "manager1")
    resp = client.get("/reports/technician-performance")
    html = resp.data.decode()
    assert "100.0%" in html  # both tasks completed -> 100% completion rate
    assert "4.0 hrs" in html  # average of 2h and 6h


def test_district_summary_maintenance_cases_counts_under_repair(app, client):
    with app.app_context():
        make_user(db, "admin", "Bugesera", "admin1")
        _seed_basic_fleet()

    login(client, "admin1")
    resp = client.get("/reports/district-summary")
    html = resp.data.decode()
    assert resp.status_code == 200
    assert "Bugesera" in html


def test_empty_filter_results_render_empty_state_not_error(app, client):
    with app.app_context():
        make_user(db, "admin", "Bugesera", "admin1")

    login(client, "admin1")
    resp = client.get("/reports/maintenance?district=NoSuchDistrict")
    assert resp.status_code == 200
    assert b"No maintenance tasks found" in resp.data


def test_pagination_splits_technician_rows_across_pages(app, client):
    with app.app_context():
        make_user(db, "admin", "Bugesera", "admin1")
        manager = make_user(db, "district_manager", "Bugesera", "manager1")
        for i in range(30):
            tech = make_user(db, "district_technician", "Bugesera", f"tech{i}")
            wp = make_water_point(db, water_point_id=f"WP-{i}", district="Bugesera")
            _seed_completed_task(manager, tech, wp)

    login(client, "admin1")
    page1 = client.get("/reports/technician-performance").data.decode()
    page2 = client.get("/reports/technician-performance?page=2").data.decode()
    assert "Page 1 of 2" in page1
    assert "Page 2 of 2" in page2
    assert page1 != page2


# --- Exports ---------------------------------------------------------------

EXPORT_ROUTES = [
    ("/reports/status/export/pdf", "status"),
    ("/reports/technician-performance/export/pdf", "technician_performance"),
    ("/reports/maintenance/export/pdf", "maintenance"),
    ("/reports/predictive-risk/export/pdf", "predictive_risk"),
    ("/reports/district-summary/export/pdf", "district_summary"),
]


def test_pdf_exports_return_valid_pdf(app, client):
    with app.app_context():
        make_user(db, "admin", "Bugesera", "admin1")
        manager = make_user(db, "district_manager", "Bugesera", "manager1")
        tech = make_user(db, "district_technician", "Bugesera", "tech1")
        wp = make_water_point(db, water_point_id="WP-1", district="Bugesera")
        _seed_completed_task(manager, tech, wp)
        _seed_basic_fleet()

    login(client, "admin1")
    for route, _ in EXPORT_ROUTES:
        resp = client.get(route)
        assert resp.status_code == 200
        assert resp.mimetype == "application/pdf"
        assert resp.data[:4] == b"%PDF"


def test_excel_exports_are_valid_workbooks_with_expected_headers(app, client):
    with app.app_context():
        make_user(db, "admin", "Bugesera", "admin1")
        manager = make_user(db, "district_manager", "Bugesera", "manager1")
        tech = make_user(db, "district_technician", "Bugesera", "tech1")
        wp = make_water_point(db, water_point_id="WP-1", district="Bugesera")
        _seed_completed_task(manager, tech, wp)
        _seed_basic_fleet()

    login(client, "admin1")
    resp = client.get("/reports/maintenance/export/excel")
    assert resp.status_code == 200
    assert resp.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    all_values = [tuple(row) for row in ws.iter_rows(values_only=True)]
    assert any(row and "Water Point" in row for row in all_values)


def test_district_summary_excel_export_handles_slash_in_title(app, client):
    """Regression test: 'District/Sector Summary Report' contains a '/' which
    openpyxl rejects as a raw worksheet title."""
    with app.app_context():
        make_user(db, "admin", "Bugesera", "admin1")
        _seed_basic_fleet()

    login(client, "admin1")
    resp = client.get("/reports/district-summary/export/excel")
    assert resp.status_code == 200
    openpyxl.load_workbook(io.BytesIO(resp.data))  # does not raise


# --- ReportLog audit trail ---------------------------------------------


def test_report_log_written_on_view_and_export(app, client):
    with app.app_context():
        make_user(db, "admin", "Bugesera", "admin1")
        _seed_basic_fleet()

    login(client, "admin1")
    client.get("/reports/status")
    client.get("/reports/status/export/pdf")
    client.get("/reports/status/export/excel")

    with app.app_context():
        logs = ReportLog.query.filter_by(report_type="status").all()
        formats = sorted(log.export_format for log in logs)
        assert formats == ["excel", "pdf", "view"]
        assert all(log.generated_by is not None for log in logs)


def test_admin_report_log_viewer_lists_entries(app, client):
    with app.app_context():
        make_user(db, "admin", "Bugesera", "admin1")
        _seed_basic_fleet()

    login(client, "admin1")
    client.get("/reports/status")
    resp = client.get("/admin/report-logs")
    assert resp.status_code == 200
    assert b"Status" in resp.data
